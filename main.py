from flask import Flask, render_template, request, send_from_directory, jsonify
from flask_cors import CORS
import os
from werkzeug.utils import secure_filename
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

app = Flask(__name__)
CORS(app)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['PROCESSED_FOLDER'] = 'processed'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls'}

# Ensure upload and processed directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['PROCESSED_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def process_timesheet(filepath):
    """Process the timesheet Excel file and generate a summarized, color-coded output."""
    try:
        # Read the Excel file, skipping the first 2 rows (metadata)
        df = pd.read_excel(filepath, skiprows=2)
        
        # Clean column names by stripping whitespace
        df.columns = df.columns.str.strip()
        
        # Convert Timesheet Period into Start Date and End Date
        df[['Start Date', 'End Date']] = df['Timesheet Period'].str.split(' - ', expand=True)
        df['Start Date'] = pd.to_datetime(df['Start Date'], format='%b %d, %Y')
        df['End Date'] = pd.to_datetime(df['End Date'], format='%b %d, %Y')
        
    except Exception as e:
        raise ValueError(f"Error reading or processing the input file: {e}")
    
    # Filter data - only "Not Submitted" status
    filtered_df = df[df['Status'].str.strip() == 'Not Submitted'].copy()
    
    if filtered_df.empty:
        raise ValueError("No 'Not Submitted' timesheets found.")
    
    # Sort and group data
    try:
        # Sort by Employee Number and Start Date
        filtered_df.sort_values(['Employee Number', 'Start Date'], inplace=True)
        
        # Group by employee and get the min/max dates
        grouped_df = filtered_df.groupby(['Employee Number', 'Employee', 'Reporting To']).agg({
            'Start Date': 'min',
            'End Date': 'max'
        }).reset_index()
        
        # Calculate duration in weeks
        grouped_df['Duration_Weeks'] = ((grouped_df['End Date'] - grouped_df['Start Date']).dt.days // 7) + 1
        
        # Format the time period string
        grouped_df['Time Period'] = grouped_df['Start Date'].dt.strftime('%b %d, %Y') + ' - ' + \
                                    grouped_df['End Date'].dt.strftime('%b %d, %Y')
        
        # Add Status column
        grouped_df['Status'] = 'Not Submitted'
        
        # Reorder columns for final output
        final_df = grouped_df[['Employee Number', 'Employee', 'Reporting To', 'Time Period', 'Status', 'Duration_Weeks']].copy()
        final_df.sort_values(by='Duration_Weeks', ascending=False, inplace=True)

        
    except Exception as e:
        raise ValueError(f"Error processing the data: {e}")
    
    # Create Excel output with color coding
    try:
        # Generate output filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f'timesheet_summary_{timestamp}.xlsx'
        output_path = os.path.join(app.config['PROCESSED_FOLDER'], output_filename)
        
        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Timesheet Summary"
        
        # Write the header row with styling
        header = ['Employee Number', 'Employee Name', 'Reporting To', 'Time Period', 'Status']
        ws.append(header)
        
        # Style the header row
        header_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Define color fills for data rows (priority order: Red > Orange > Yellow > White)
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        orange_fill = PatternFill(start_color='FFA500FF', end_color='FFA500FF', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
        
        # Write data rows and apply color coding to all cells in each row
        for _, row in final_df.iterrows():
            # Determine the fill color based on duration
            duration = row['Duration_Weeks']
            if duration > 3:
                fill = red_fill
            elif duration == 3:
                fill = orange_fill
            elif duration == 2:
                fill = yellow_fill
            else:
                fill = white_fill
            
            # Add the row data
            ws.append([
                row['Employee Number'],
                row['Employee'],
                row['Reporting To'],
                row['Time Period'],
                row['Status']
            ])
            
            # Apply color to all cells in the current row
            for cell in ws[ws.max_row]:
                cell.fill = fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(output_path)
        return output_filename
        
    except Exception as e:
        raise ValueError(f"Error creating the output file: {e}")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(upload_path)
        
        try:
            processed_filename = process_timesheet(upload_path)
            return jsonify({
                'success': True,
                'filename': processed_filename
            })
        except ValueError as e:
            return jsonify({'success': False, 'error': str(e)}), 400
        except Exception as e:
            return jsonify({'success': False, 'error': 'An unexpected error occurred'}), 500
    else:
        return jsonify({
            'success': False,
            'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'
        }), 400

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(
        app.config['PROCESSED_FOLDER'],
        filename,
        as_attachment=True
    )

if __name__ == '__main__':
    app.run(debug=True)