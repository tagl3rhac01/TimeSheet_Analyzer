import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import os

def process_timesheet(input_file_path, output_file_path):
    """
    Process the timesheet Excel file and generate a summarized, color-coded output.
    
    Args:
        input_file_path (str): Path to the input Excel file
        output_file_path (str): Path where the output Excel file will be saved
    """
    
    # Step 1: Read and clean data
    try:
        # Read the Excel file, skipping the first 2 rows (metadata)
        df = pd.read_excel(input_file_path, skiprows=2)
        
        # Clean column names by stripping whitespace
        df.columns = df.columns.str.strip()
        
        # Convert Timesheet Period into Start Date and End Date
        df[['Start Date', 'End Date']] = df['Timesheet Period'].str.split(' - ', expand=True)
        df['Start Date'] = pd.to_datetime(df['Start Date'], format='%b %d, %Y')
        df['End Date'] = pd.to_datetime(df['End Date'], format='%b %d, %Y')
        
    except Exception as e:
        print(f"Error reading or processing the input file: {e}")
        return False
    
    # Step 2: Filter data - only "Not Submitted" status
    filtered_df = df[df['Status'].str.strip() == 'Not Submitted'].copy()
    
    if filtered_df.empty:
        print("No 'Not Submitted' timesheets found.")
        return False
    
    # Step 3: Sort and group data
    try:
        # Sort by Employee Number and Start Date
        filtered_df.sort_values(['Employee Number', 'Start Date'], inplace=True)
        
        # Group by employee and get the min/max dates
        grouped_df = filtered_df.groupby(['Employee Number', 'Employee', 'Reporting To']).agg({
            'Start Date': 'min',
            'End Date': 'max'
        }).reset_index()
        
        # Step 4: Calculate duration in weeks
        grouped_df['Duration_Weeks'] = ((grouped_df['End Date'] - grouped_df['Start Date']).dt.days // 7) + 1
        
        # Format the time period string
        grouped_df['Time Period'] = grouped_df['Start Date'].dt.strftime('%b %d, %Y') + ' - ' + \
                                    grouped_df['End Date'].dt.strftime('%b %d, %Y')
        
        # Add Status column
        grouped_df['Status'] = 'Not Submitted'
        
        # Reorder columns for final output
        final_df = grouped_df[['Employee Number', 'Employee', 'Reporting To', 'Time Period', 'Status', 'Duration_Weeks']].copy()
        final_df.sort_values(by='Duration_Weeks', ascending=False, inplace=True)

        
    except Exception as e:
        print(f"Error processing the data: {e}")
        return False
    
    # Step 5 & 6: Create Excel output with color coding
    try:
        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Timesheet Summary"
        
        # Write the header row with styling
        header = ['Employee Number', 'Employee Name', 'Reporting To', 'Time Period', 'Status']
        ws.append(header)
        
        # Style the header row
        header_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')  # Light gray
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Define color fills for data rows (updated priority order)
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')  # Red (highest priority)
        orange_fill = PatternFill(start_color='FFA500FF', end_color='FFA500FF', fill_type='solid')  # Orange
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')  # Yellow
        white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')  # White (lowest priority)
        
        # Write data rows and apply color coding to all cells in each row
        for _, row in final_df.iterrows():
            # Determine the fill color based on duration (updated priority order)
            duration = row['Duration_Weeks']
            if duration > 3:
                fill = red_fill
            elif duration == 3:
                fill = orange_fill
            elif duration == 2:
                fill = yellow_fill
            else:
                fill = white_fill
            
            # Add the row data (excluding Duration_Weeks which is only for internal use)
            ws.append([
                row['Employee Number'],
                row['Employee'],
                row['Reporting To'],
                row['Time Period'],
                row['Status']
            ])
            
            # Apply color to all cells in the current row
            for cell in ws[ws.max_row]:
                cell.fill = fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(output_file_path)
        print(f"Successfully created output file: {output_file_path}")
        return True
        
    except Exception as e:
        print(f"Error creating the output file: {e}")
        return False

if __name__ == "__main__":
    # Hardcoded file paths (you can modify these)
    input_file = r"C:\Users\Jatin Chandok\Desktop\INTERNAL TASK\Timesheets Status Report in Duration- The Icon Group (15).xlsx"  # Replace with your input file name
    output_file = "summary_colored.xlsx"
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"Input file not found: {input_file}")
        print("Please ensure the input file exists in the same directory as this script.")
    else:
        # Process the timesheet
        process_timesheet(input_file, output_file)